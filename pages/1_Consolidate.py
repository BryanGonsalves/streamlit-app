from copy import copy
from io import BytesIO
from typing import Dict, List, Optional, Sequence, Tuple

import openpyxl
import streamlit as st
from openpyxl.worksheet.worksheet import Worksheet

from filesplit import get_column_letter_by_header


def _load_uploaded_files(uploaded_files) -> Sequence[Tuple[str, bytes]]:
    """Extract byte content from uploaded files once."""
    buffers: List[Tuple[str, bytes]] = []
    for uploaded in uploaded_files:
        buffers.append((uploaded.name, uploaded.getvalue()))
    return tuple(buffers)


def _last_data_row(ws: Worksheet, header_row: int) -> int:
    """Find the last row containing data below the header."""
    for row_idx in range(ws.max_row, header_row, -1):
        for cell in ws[row_idx]:
            if cell.value not in (None, ""):
                return row_idx
    return header_row


def _append_row_with_style(destination_ws: Worksheet, source_row) -> None:
    """Append a row to destination_ws copying both values and styles."""
    values = [cell.value for cell in source_row]
    destination_ws.append(values)
    dest_row_idx = destination_ws.max_row
    for dest_cell, src_cell in zip(destination_ws[dest_row_idx], source_row):
        if src_cell.has_style:
            dest_cell.font = copy(src_cell.font)
            dest_cell.border = copy(src_cell.border)
            dest_cell.fill = copy(src_cell.fill)
            dest_cell.number_format = copy(src_cell.number_format)
            dest_cell.protection = copy(src_cell.protection)
            dest_cell.alignment = copy(src_cell.alignment)
        dest_cell.hyperlink = src_cell.hyperlink
        dest_cell.comment = copy(src_cell.comment) if src_cell.comment else None


@st.cache_data(show_spinner=False)
def build_consolidated_workbook(
    uploaded_files: Sequence[Tuple[str, bytes]], header_label: str
) -> Tuple[Optional[bytes], Dict[str, List[str]], int]:
    """
    Merge the provided workbooks into a single workbook.

    Returns the combined workbook bytes, any sheets missing the target header,
    and the total merged row count.
    """
    if not uploaded_files:
        return None, {}, 0

    header_variants = [header_label, f"{header_label}s"]
    missing_by_file: Dict[str, List[str]] = {}

    base_name, base_bytes = uploaded_files[0]
    base_wb = openpyxl.load_workbook(BytesIO(base_bytes))
    sheet_header_rows: Dict[str, int] = {}

    for sheet_name in base_wb.sheetnames:
        ws = base_wb[sheet_name]
        col_letter, header_row = get_column_letter_by_header(ws, header_variants)
        if not col_letter:
            missing_by_file.setdefault(base_name, []).append(sheet_name)
            continue
        sheet_header_rows[sheet_name] = header_row

    total_rows = 0

    # Count rows already present in the base workbook
    for sheet_name, header_row in sheet_header_rows.items():
        ws = base_wb[sheet_name]
        total_rows += max(0, _last_data_row(ws, header_row) - header_row)

    # Process additional workbooks
    for filename, file_bytes in uploaded_files[1:]:
        workbook = openpyxl.load_workbook(BytesIO(file_bytes))
        for sheet_name in workbook.sheetnames:
            ws_src = workbook[sheet_name]
            col_letter, header_row = get_column_letter_by_header(ws_src, header_variants)
            if not col_letter:
                missing_by_file.setdefault(filename, []).append(sheet_name)
                continue

            if sheet_name not in sheet_header_rows:
                # Create the sheet in the base workbook copying header row and styles.
                ws_dest = base_wb.create_sheet(title=sheet_name)
                sheet_header_rows[sheet_name] = header_row
                header_row_cells = ws_src[header_row]
                _append_row_with_style(ws_dest, header_row_cells)
                # Ensure rows below header start empty
                ws_dest.delete_rows(sheet_header_rows[sheet_name] + 1, ws_dest.max_row - sheet_header_rows[sheet_name])
            else:
                ws_dest = base_wb[sheet_name]

            dest_header_row = sheet_header_rows[sheet_name]
            last_row = _last_data_row(ws_dest, dest_header_row)

            for row_idx in range(header_row + 1, ws_src.max_row + 1):
                source_row = ws_src[row_idx]
                if all(cell.value in (None, "") for cell in source_row):
                    continue
                _append_row_with_style(ws_dest, source_row)
                total_rows += 1

            # Ensure we preserve spacing: remove unintended blank rows between header and data
            if last_row == dest_header_row and ws_dest.max_row > dest_header_row:
                pass  # header only previously, nothing to clean
        workbook.close()

    buffer = BytesIO()
    base_wb.save(buffer)
    base_wb.close()

    return buffer.getvalue(), missing_by_file, total_rows


def main():
    st.title("Consolidate")
    st.write(
        "Combine multiple workbooks into a single file keyed by Team Lead or Mentor. "
        "Perfect for recreating a master workbook from individual exports."
    )

    st.sidebar.header("How to use")
    st.sidebar.write(
        "1. Choose whether to consolidate by Team Lead or Mentor and provide an optional output name.\n"
        "2. Upload one or more workbooks structured identically.\n"
        "3. Press **Consolidate files** to merge the data.\n"
        "4. Download the single combined workbook."
    )

    default_filter = st.session_state.get("consolidate_filter_by_mentor", False)
    toggle_label = f"Consolidate by {'Mentor' if default_filter else 'Team Lead'}"
    filter_by_mentor = st.toggle(toggle_label, value=default_filter, key="consolidate_filter_by_mentor")
    target_header = "Mentor" if filter_by_mentor else "Team Lead"

    output_name_input = st.text_input(
        "Output workbook name",
        value="",
        placeholder="e.g., Consolidated September 2025",
        help="The generated file name will append .xlsx automatically if you omit it.",
        key="consolidate_output_name",
    )

    uploaded_files = st.file_uploader(
        "Upload one or more workbooks (.xlsx) to consolidate", type=["xlsx"], accept_multiple_files=True
    )

    file_buffers: Optional[Sequence[Tuple[str, bytes]]] = None
    result_key = None
    if uploaded_files:
        file_buffers = _load_uploaded_files(uploaded_files)
        result_key = (
            file_buffers,
            target_header,
            output_name_input.strip(),
        )

    stored_results = st.session_state.get("consolidation_results")
    should_show_results = stored_results is not None and stored_results.get("key") == result_key

    submitted = st.button("Consolidate files", use_container_width=True)

    if submitted:
        if not file_buffers:
            st.error("Upload at least one workbook before consolidating.")
            st.session_state.pop("consolidation_results", None)
            return

        with st.spinner("Consolidating workbooks..."):
            workbook_bytes, missing_sheets, row_count = build_consolidated_workbook(
                file_buffers, target_header
            )

        if not workbook_bytes:
            st.error(
                f"No {target_header.lower()} data was found across the uploaded workbooks. "
                "Confirm the selected column header exists."
            )
            st.session_state.pop("consolidation_results", None)
            return

        stored_results = {
            "workbook": workbook_bytes,
            "missing_sheets": missing_sheets,
            "row_count": row_count,
            "target_header": target_header,
            "key": result_key,
        }
        st.session_state["consolidation_results"] = stored_results
    elif should_show_results:
        stored_results = st.session_state["consolidation_results"]
    else:
        if not uploaded_files:
            st.session_state.pop("consolidation_results", None)
            st.info("Start by uploading the workbooks you want to merge.")
        else:
            st.info("Click **Consolidate files** to merge the uploaded workbooks.")
            if stored_results and stored_results.get("key") != result_key:
                st.warning("Inputs changed. Click **Consolidate files** to refresh.")
        return

    missing_sheets = stored_results.get("missing_sheets", {})
    if missing_sheets:
        missing_messages = [
            f"- **{filename}**: {', '.join(sheets)}" for filename, sheets in missing_sheets.items()
        ]
        st.warning(
            "Some uploaded workbooks did not contain the selected column and were skipped:\n"
            + "\n".join(missing_messages)
        )

    row_count = stored_results.get("row_count", 0)
    st.success(
        f"Merged {row_count} row(s) across all sheets "
        f"for {target_header.lower()} consolidation."
    )

    output_name = output_name_input.strip() or f"consolidated-{target_header.lower().replace(' ', '-')}"
    if not output_name.lower().endswith(".xlsx"):
        output_name += ".xlsx"

    st.download_button(
        label=f"Download {output_name}",
        data=stored_results["workbook"],
        file_name=output_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
