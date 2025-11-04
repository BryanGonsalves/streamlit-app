import zipfile
from io import BytesIO
from typing import Dict, List, Optional, Set, Tuple

import openpyxl
import streamlit as st

CANONICAL_LEAD_ALIASES = {
    "fauziahasansiddiqui": "Fauzia Hasan",
}


def _normalize_header(value) -> str:
    """Lowercase header name stripped of surrounding whitespace."""
    return "".join(str(value).split()).lower()


def _canonicalize_lead(name) -> str:
    """Return canonical team lead name so aliases collapse into one."""
    if name is None:
        return ""
    stripped = str(name).strip()
    if not stripped:
        return ""

    normalized = _normalize_header(stripped)
    return CANONICAL_LEAD_ALIASES.get(normalized, stripped)


def sanitize_prefix(prefix: str) -> str:
    """Ensure prefix ends with a single space when provided."""
    if not prefix:
        return ""
    return prefix if prefix.endswith(" ") else f"{prefix} "


def get_column_letter_by_header(sheet, header_names) -> Tuple[Optional[str], Optional[int]]:
    """Return the column letter and header row index for the first matching header."""
    if isinstance(header_names, str):
        header_names = [header_names]

    normalized_targets = {_normalize_header(name) for name in header_names if name}

    # Search the first few rows for headers because some sheets have blank top rows.
    max_scan_row = min(sheet.max_row, 10) or 1
    for row in sheet.iter_rows(min_row=1, max_row=max_scan_row):
        for cell in row:
            if cell.value is None:
                continue

            normalized_value = _normalize_header(cell.value)
            if not normalized_value:
                continue

            for normalized_target in normalized_targets:
                # Accept exact matches and simple plural forms (Team Lead vs Team Leads).
                if (
                    normalized_value == normalized_target
                    or normalized_value.rstrip("s") == normalized_target.rstrip("s")
                ):
                    return cell.column_letter, cell.row

    return None, None


@st.cache_data(show_spinner=False)
def generate_entity_workbooks(
    master_file_bytes: bytes, header_label: str
) -> Tuple[List[str], Dict[str, bytes], List[str]]:
    """Return the sorted entity names, workbook bytes per entity, and sheets lacking the column."""
    original_wb = openpyxl.load_workbook(BytesIO(master_file_bytes))
    sheet_names = original_wb.sheetnames

    entity_names: Set[str] = set()
    missing_sheets: List[str] = []

    header_variants = [header_label, f"{header_label}s"]

    for sheet_name in sheet_names:
        ws = original_wb[sheet_name]
        col_letter, header_row = get_column_letter_by_header(ws, header_variants)
        if not col_letter:
            missing_sheets.append(sheet_name)
            continue

        for row in range(header_row + 1, ws.max_row + 1):  # Skip header row
            val = ws[f"{col_letter}{row}"].value
            if val is None:
                continue

            lead_name = _canonicalize_lead(val)
            if lead_name:
                entity_names.add(lead_name)

    sorted_entities = sorted(entity_names)
    original_wb.close()

    if not sorted_entities:
        return [], {}, missing_sheets

    workbooks: Dict[str, bytes] = {}
    for lead in sorted_entities:
        wb_copy = openpyxl.load_workbook(BytesIO(master_file_bytes))

        for sheet_name in sheet_names:
            ws_copy = wb_copy[sheet_name]
            col_letter, header_row = get_column_letter_by_header(ws_copy, header_variants)
            if not col_letter:
                continue  # Skip sheet if no target column

            # Loop bottom-up to delete rows not matching this lead
            rows_to_delete = []
            start = None
            length = 0

            for row in range(header_row + 1, ws_copy.max_row + 1):
                cell_value = ws_copy[f"{col_letter}{row}"].value
                cell_lead = _canonicalize_lead(cell_value)
                if cell_lead == lead:
                    if start is not None:
                        rows_to_delete.append((start, length))
                        start = None
                        length = 0
                    continue

                if start is None:
                    start = row
                    length = 1
                else:
                    if row == start + length:
                        length += 1
                    else:
                        rows_to_delete.append((start, length))
                        start = row
                        length = 1

            if start is not None and length:
                rows_to_delete.append((start, length))

            for row_start, block_length in reversed(rows_to_delete):
                ws_copy.delete_rows(row_start, block_length)

        buffer = BytesIO()
        wb_copy.save(buffer)
        wb_copy.close()
        workbooks[lead] = buffer.getvalue()

    return sorted_entities, workbooks, missing_sheets


def _create_zip_from_workbooks(workbooks: Dict[str, bytes], prefix: str) -> bytes:
    """Package generated workbooks into a zip archive."""
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w") as zip_file:
        for lead, workbook_bytes in workbooks.items():
            filename = f"{prefix}{lead}.xlsx" if prefix else f"{lead}.xlsx"
            zip_file.writestr(filename, workbook_bytes)
    buffer.seek(0)
    return buffer.getvalue()


def main() -> None:
    st.set_page_config(page_title="File Splitter", page_icon="📄", layout="centered")
    st.title("File Splitter")
    st.write("Enter your consolidated workbook to automatically generate individual files for each selected role.")

    st.sidebar.header("How to use")
    st.sidebar.write(
        "1. Pick the role to split by and, if desired, set a filename prefix.\n"
        "2. Upload the consolidated workbook (header labels in row 1).\n"
        "3. Press **Generate files** to process the workbook.\n"
        "4. Review the detected names and download the outputs."
    )

    default_filter = st.session_state.get("filter_by_mentor", False)
    toggle_label = f"Filter by {'Mentor' if default_filter else 'Team Lead'}"
    filter_by_mentor = st.toggle(toggle_label, value=default_filter, key="filter_by_mentor")
    target_header = "Mentor" if filter_by_mentor else "Team Lead"
    prefix_input = st.text_input(
        "Optional filename prefix",
        value="",
        placeholder="e.g., Student Master 2025_Team ",
        help="Prefix added ahead of each generated filename.",
        key="split_prefix_input",
    )
    sanitized_prefix = sanitize_prefix(prefix_input)

    uploaded_file = st.file_uploader("Upload the consolidated workbook (.xlsx)", type=["xlsx"])

    master_bytes = uploaded_file.getvalue() if uploaded_file is not None else None
    result_key = None
    if master_bytes is not None:
        result_key = (
            uploaded_file.name,
            len(master_bytes),
            target_header,
            sanitized_prefix,
        )

    stored_results = st.session_state.get("split_results")
    should_show_results = stored_results is not None and stored_results.get("key") == result_key

    submitted = st.button("Generate files", use_container_width=True)

    if submitted:
        if master_bytes is None:
            st.error("Upload an Excel workbook before generating files.")
            st.session_state.pop("split_results", None)
            return
        if not master_bytes:
            st.error("The uploaded file appears to be empty.")
            st.session_state.pop("split_results", None)
            return

        with st.spinner("Processing workbook..."):
            leads, workbooks, missing_sheets = generate_entity_workbooks(master_bytes, target_header)

        if not leads:
            st.error(f"No {target_header.lower()}s were found in the uploaded workbook.")
            st.session_state.pop("split_results", None)
            return

        stored_results = {
            "leads": leads,
            "workbooks": workbooks,
            "missing_sheets": missing_sheets,
            "target_header": target_header,
            "prefix": sanitized_prefix,
            "key": result_key,
        }
        st.session_state["split_results"] = stored_results
        sanitized_prefix = stored_results["prefix"]
    elif should_show_results:
        stored_results = st.session_state["split_results"]
        leads = stored_results["leads"]
        workbooks = stored_results["workbooks"]
        missing_sheets = stored_results["missing_sheets"]
        target_header = stored_results["target_header"]
        sanitized_prefix = stored_results["prefix"]
    else:
        if uploaded_file is None:
            st.session_state.pop("split_results", None)
            st.info("Start by uploading an Excel workbook.")
        else:
            st.info("Click **Generate files** to process the workbook.")
            if stored_results and stored_results.get("key") != result_key:
                st.warning("Inputs changed. Click **Generate files** to refresh.")
        return

    if missing_sheets:
        st.warning(
            f"The following sheets do not contain a '{target_header}' column and were skipped: "
            + ", ".join(missing_sheets)
        )

    if not leads:
        st.error(f"No {target_header.lower()}s were found in the uploaded workbook.")
        return

    st.success(f"Found {len(leads)} {target_header.lower()}s.")
    st.write("Download the generated files below.")

    zip_bytes = _create_zip_from_workbooks(workbooks, sanitized_prefix)
    st.download_button(
        label="Download all workbooks as ZIP",
        data=zip_bytes,
        file_name=f"{target_header.lower().replace(' ', '-')}-workbooks.zip",
        mime="application/zip",
        use_container_width=True,
    )

    st.subheader("Individual downloads")
    columns = st.columns(2)
    for index, lead in enumerate(leads):
        column = columns[index % len(columns)]
        column.download_button(
            label=f"Download {lead}",
            data=workbooks[lead],
            file_name=f"{sanitized_prefix}{lead}.xlsx" if sanitized_prefix else f"{lead}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_{lead}",
            use_container_width=True,
        )


if __name__ == "__main__":
    main()
